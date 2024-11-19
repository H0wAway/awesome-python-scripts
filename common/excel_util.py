import os
import re
import time

import openpyxl
import pandas as pd
import xlwings as xw


def excel_column_to_number(column_str):
    """
    将 Excel 列号字符串转换为整数。

    Args:
        column_str (str): Excel 列号字符串，例如 "A" 或 "AA"。

    Example:
        excel_column_to_number("A")  # 1
        excel_column_to_number("AA")  # 27

    Returns:
        int: 对应的列号整数值。
    """
    number = 0
    for char in column_str:
        number = number * 26 + (ord(char) - ord('A') + 1)
    return number


def number_to_excel_column(number):
    """
    将整数转换为 Excel 列号字符串。

    Args:
        number (int): 列号整数，例如 1 或 27。

    Returns:
        str: 对应的 Excel 列号字符串。
    """
    column_str = ""
    while number > 0:
        number, remainder = divmod(number - 1, 26)
        column_str = chr(remainder + ord('A')) + column_str
    return column_str


def unmerge_and_fill_cells(worksheet):
    """
    拆分所有的合并单元格，并赋予合并之前的值。（由于openpyxl并没有提供拆分并填充的方法，所以使用该方法进行完成）
    :param worksheet:
    :return:
    """
    all_merged_cell_ranges = list(worksheet.merged_cells.ranges)

    for merged_cell_range in all_merged_cell_ranges:
        merged_cell = merged_cell_range.start_cell
        worksheet.unmerge_cells(range_string=merged_cell_range.coord)

        for row_index, col_index in merged_cell_range.cells:
            cell = worksheet.cell(row=row_index, column=col_index)
            cell.value = merged_cell.value


def unmerge_cell(filename):
    """
    读取原始xlsx文件，拆分并填充单元格，然后生成临时文件
    Example：
        unmerge_cell(r"C:\123——副本.xlsx")
    :param filename:  文件路径
    :return:  生成的临时文件路径 例如：C:\123——副本_temp.xlsx
    """
    wb = openpyxl.load_workbook(filename)
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        unmerge_and_fill_cells(sheet)
    filename = filename.replace(".xls", "_temp.xls")
    wb.save(filename)
    wb.close()

    # openpyxl保存之后，再用pandas读取会存在公式无法读取到的情况，使用下面方式就可以了（如果excel不涉及公式，可以删除下面内容）
    # 原理为：使用windows打开excel，然后另存为一下
    # from win32com.client import Dispatch
    #
    # xlApp = Dispatch("Excel.Application")
    # xlApp.Visible = False
    # xlBook = xlApp.Workbooks.Open(str(Path(".").absolute() / filename))  # 这里必须填绝对路径
    # xlBook.Save()
    # xlBook.Close()

    return filename


def read_excel_column_value(filename: str, sheet_name: str, column_index: int) -> list[str]:
    """
    使用 openpyxl 读取excel文件的指定sheet的指定列的数据，返回 list[str]
    :rtype: object
    :param filename: 文件路径
    :param sheet_name: sheet名称
    :param column_index: 列索引
    :return: 读取文本数据，类型为list[str]
    """
    # data_only=True 读取的是单元格的最终值，而不是公式
    wb = openpyxl.load_workbook(filename, data_only=True)
    sheet = wb[sheet_name]
    column_data = []
    for row in sheet.iter_rows(min_row=2, min_col=column_index, max_col=column_index):
        for cell in row:
            column_data.append(cell.value)
    wb.close()
    return column_data


def read_excel_to_pandas(filename: str, sheet_name: str) -> pd.DataFrame:
    """
    使用 openpyxl 读取excel文件的指定sheet的数据，返回 pd.DataFrame
    :param filename: 文件路径
    :param sheet_name: sheet名称
    :return: 读取文本数据，类型为pd.DataFrame
    """
    # data_only=True 读取的是单元格的最终值，而不是公式
    wb = openpyxl.load_workbook(filename, data_only=True)
    sheet = wb[sheet_name]
    data = sheet.values
    df = pd.DataFrame(data)
    wb.close()
    return df


def xlwings2openpyxl():
    """
    xlwings -> openpyxl无格式粘贴
    :return:
    """
    # 原始文件路径
    original_file_path = r'C:\Users\haowei\Desktop\固德威wSolar平台逆变器错误代码对照表.xlsx'
    # 备份文件路径
    backup_file_path = 'E:\\BaiduSyncdisk\\temp.xlsx'
    # 检查原始文件是否存在
    if not os.path.exists(original_file_path):
        print(f"Error: The file {original_file_path} does not exist.")
    if os.path.exists(backup_file_path):
        os.remove(backup_file_path)
    original_wb = xw.Book(original_file_path)
    time.sleep(3)
    # 创建备份文件
    backup_wb = openpyxl.Workbook()
    # 移除创建空工作簿后自动生成的工作表（可选）
    backup_wb.remove(backup_wb.active)
    for sheet in original_wb.sheets:
        # 在备份工作簿中创建对应的工作表
        backup_sheet = backup_wb.create_sheet(title=sheet.name)
        # 读取已用范围
        end_cell_indexs = re.search(r":\$([A-Z]*)\$(\d+)", sheet.used_range.address).groups()
        end_cell_col_index = int(excel_column_to_number(end_cell_indexs[0]))
        end_cell_row_index = int(end_cell_indexs[1])
        # 读取原始工作表的所有数据
        data = sheet.used_range.value
        # 将数据写入备份工作表
        for row_index, row_data in enumerate(data, start=1):
            for col_index, cell_value in enumerate(row_data, start=1):
                # 如果单元格没有数据，则跳过
                if cell_value is None:
                    continue
                backup_sheet.cell(row=row_index, column=col_index, value=cell_value)
        # 处理合并单元格
        #         backup_sheet.merge_cells(
        #             start_row=start_cell.Row,
        #             start_column=start_cell.Column,
        #             end_row=end_cell.Row,
        #             end_column=end_cell.Column,
        #         )
    # 保存备份文件
    backup_wb.save(backup_file_path)
    time.sleep(2)
    # 关闭原始文件
    original_wb.close()
